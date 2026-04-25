"""
All DRF API views for REMS. Views are thin; business logic lives in services.py.
"""

from django.utils import timezone
from django.db.models import Q
from rest_framework import status, viewsets, generics, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.exceptions import TokenError
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
import csv
import io
from datetime import date, timedelta

from .models import (
    User, Department, AttendancePolicy, Attendance,
    WorkSession, BreakSession, IdleLog, LeaveRequest, Holiday, AuditLog, OTPRecord, Notification,
    ScreenCapture
)
from .serializers import (
    UserListSerializer, UserDetailSerializer, UserCreateSerializer,
    UserUpdateSerializer, LoginSerializer, ChangePasswordSerializer,
    DepartmentSerializer, AttendancePolicySerializer, AttendanceSerializer,
    WorkSessionSerializer, BreakSessionSerializer, IdleLogSerializer,
    LeaveRequestSerializer, LeaveApplySerializer, LeaveReviewSerializer,
    HolidaySerializer, AuditLogSerializer, AttendanceReviewSerializer, NotificationSerializer,
    ScreenCaptureSerializer,
)
from .permissions import IsAdmin, IsManager, IsEmployee, IsOwnerOrManager
from .services import (
    AttendanceService, WorkSessionService, BreakSessionService,
    IdleService, StatusService, LeaveService, ReportService, AuditService,
    NotificationService
)


# ═══════════════════════════════════════════════════════════════════════════════
# Module 1: AUTH & SECURITY
# ═══════════════════════════════════════════════════════════════════════════════

import random
from django.core.mail import send_mail
from django.conf import settings

class RequestOTPView(APIView):
    """Step 1: Request an OTP for Login or Registration"""
    permission_classes = [AllowAny]

    def post(self, request):
        email = request.data.get('email')
        if not email:
            return Response({'detail': 'Email is required.'}, status=status.HTTP_400_BAD_REQUEST)
        
        # In a real app we'd throttle this.
        # Generate 6 digit OTP
        otp_code = f"{random.randint(100000, 999999)}"
        
        # Save to DB
        OTPRecord.objects.create(user_email=email, otp_code=otp_code)

        # Send Email (Console backend will print this)
        send_mail(
            'Your REMS Verification Code',
            f'Your OTP code is: {otp_code}',
            settings.DEFAULT_FROM_EMAIL if hasattr(settings, 'DEFAULT_FROM_EMAIL') else 'noreply@rems.com',
            [email],
            fail_silently=False,
        )

        # Determine if user exists for frontend logic
        user_exists = User.objects.filter(email=email).exists()
        
        return Response({
            'detail': 'OTP sent successfully.',
            'user_exists': user_exists,
            'otp': otp_code
        })

class VerifyOTPView(APIView):
    """Step 2 (Registration): Verify OTP before collecting profile details"""
    permission_classes = [AllowAny]

    def post(self, request):
        email = request.data.get('email')
        otp_code = request.data.get('otp')

        if not email or not otp_code:
            return Response({'detail': 'Email and OTP are required.'}, status=status.HTTP_400_BAD_REQUEST)
        
        record = OTPRecord.objects.filter(user_email=email, otp_code=otp_code, is_verified=False).first()
        if not record:
            return Response({'detail': 'Invalid or expired OTP.'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Check expiration (e.g. 10 mins)
        if (timezone.now() - record.created_at).total_seconds() > 600:
            return Response({'detail': 'OTP expired.'}, status=status.HTTP_400_BAD_REQUEST)

        record.is_verified = True
        record.save()
        
        return Response({'detail': 'OTP verified successfully.', 'verified_token': otp_code})


class RegisterProfileView(APIView):
    """Step 3 (Registration): Submit User Details + Verified OTP to finish signup"""
    permission_classes = [AllowAny]

    def post(self, request):
        email = request.data.get('email')
        otp_code = request.data.get('verified_token')
        first_name = request.data.get('first_name')
        last_name = request.data.get('last_name')
        password = request.data.get('password')

        if not all([email, otp_code, first_name, last_name, password]):
            return Response({'detail': 'All fields are required.'}, status=status.HTTP_400_BAD_REQUEST)

        # Check OTP was actually verified recently
        record = OTPRecord.objects.filter(user_email=email, otp_code=otp_code, is_verified=True).first()
        if not record:
            return Response({'detail': 'Invalid verification session.'}, status=status.HTTP_400_BAD_REQUEST)

        if User.objects.filter(email=email).exists():
            return Response({'detail': 'User already exists.'}, status=status.HTTP_400_BAD_REQUEST)

        # Create user
        user = User.objects.create_user(
            email=email,
            password=password,
            first_name=first_name,
            last_name=last_name,
            role='employee', # Default to employee
            is_active=True
        )

        # Invalidate OTP
        record.delete()
        
        # Notify Admins about new registration
        NotificationService.notify_based_on_role(
            user,
            "New User Registered",
            f"A new user {user.full_name} ({user.email}) has registered and is awaiting assignment.",
            "system"
        )
        
        return Response({'detail': 'Registration complete. You can now log in.'}, status=status.HTTP_201_CREATED)


class LoginView(APIView):
    """JWT login with 2FA OTP verification required."""
    permission_classes = [AllowAny]

    def post(self, request):
        # We expect email, password, AND otp for login flow.
        email = request.data.get('email')
        password = request.data.get('password')
        otp_code = request.data.get('otp')

        if not all([email, password, otp_code]):
            return Response({'detail': 'Email, password, and OTP are required for login.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            return Response({'detail': 'Invalid credentials.'}, status=status.HTTP_401_UNAUTHORIZED)

        if not user.is_active:
            return Response({'detail': 'Account is inactive.'}, status=status.HTTP_403_FORBIDDEN)

        if not user.check_password(password):
            return Response({'detail': 'Invalid credentials.'}, status=status.HTTP_401_UNAUTHORIZED)

        # Verify OTP
        record = OTPRecord.objects.filter(user_email=email, otp_code=otp_code, is_verified=False).first()
        if not record:
            return Response({'detail': 'Invalid or expired OTP.'}, status=status.HTTP_401_UNAUTHORIZED)
        
        # Check expiration (e.g. 10 mins)
        if (timezone.now() - record.created_at).total_seconds() > 600:
            return Response({'detail': 'OTP expired.'}, status=status.HTTP_400_BAD_REQUEST)

        record.is_verified = True
        record.save()
        # Clean up
        OTPRecord.objects.filter(user_email=email).delete()

        # Issue JWT tokens
        refresh = RefreshToken.for_user(user)

        # Module 4: Auto-create attendance on login
        attendance, created = AttendanceService.get_or_create_today(user)

        # Audit log
        AuditService.log(user, 'login', f'User {user.email} logged in with 2FA', request)

        # Notify managers/admins of employee login
        NotificationService.notify_based_on_role(user, "User Login", f"{user.full_name} has logged in.", "system")

        return Response({
            'access': str(refresh.access_token),
            'refresh': str(refresh),
            'user': {
                'id': str(user.id),
                'email': user.email,
                'full_name': user.full_name,
                'role': user.role,
                'department': str(user.department_id) if user.department_id else None,
                'department_name': user.department.name if user.department else None,
            },
            'attendance_created': created,
        })


class LogoutView(APIView):
    """Blacklist the refresh token on logout."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            from django.core.cache import cache
            user = request.user
            
            refresh_token = request.data.get('refresh')
            if refresh_token:
                token = RefreshToken(refresh_token)
                token.blacklist()
            
            # 1. Clear presence cache for instant offline status
            cache.delete(f'presence_{user.id}')
            
            # 2. Broadcast status change immediately
            StatusService.broadcast_status_change(user)
            
            # 3. Stop any active work/idle sessions upon logout for accurate time tracking
            WorkSessionService.stop_session(user)
            
            # 4. Notify managers/admins of employee logout
            NotificationService.notify_based_on_role(user, "User Logout", f"{user.full_name} has logged out.", "system")
            
            AuditService.log(user, 'logout', f'User {user.email} logged out', request)
            return Response({'detail': 'Logged out successfully.'})
        except TokenError:
            return Response({'detail': 'Invalid or expired token.'}, status=status.HTTP_400_BAD_REQUEST)


class ChangePasswordView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = ChangePasswordSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        request.user.set_password(serializer.validated_data['new_password'])
        request.user.save()
        return Response({'detail': 'Password changed successfully.'})


class MeView(APIView):
    """Return current user profile."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        serializer = UserDetailSerializer(request.user)
        return Response(serializer.data)

    def patch(self, request):
        serializer = UserUpdateSerializer(request.user, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(UserDetailSerializer(request.user).data)


# ═══════════════════════════════════════════════════════════════════════════════
# Module 2 & 3: USER MANAGEMENT + ROLE AUTHORIZATION
# ═══════════════════════════════════════════════════════════════════════════════

class UserViewSet(viewsets.ModelViewSet):
    """Full CRUD for users. Admin only for write; Manager read-only."""
    queryset = User.objects.select_related('department', 'manager').all()
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['role', 'department', 'is_active']
    search_fields = ['email', 'first_name', 'last_name']
    ordering_fields = ['created_at', 'email', 'first_name']
    ordering = ['-created_at']

    def get_permissions(self):
        if self.action in ('list', 'retrieve'):
            return [IsManager()]
        return [IsAdmin()]

    def get_serializer_class(self):
        if self.action == 'create':
            return UserCreateSerializer
        if self.action in ('update', 'partial_update'):
            return UserUpdateSerializer
        if self.action == 'list':
            return UserListSerializer
        return UserDetailSerializer

    def get_queryset(self):
        user = self.request.user
        if user.role == 'manager':
            # Managers can only see their subordinates
            return User.objects.filter(
                Q(manager=user) | Q(id=user.id)
            ).select_related('department', 'manager')
        return super().get_queryset()

    def perform_create(self, serializer):
        user = serializer.save()
        # Notify superiors (Manager & Admins) about the new user
        NotificationService.notify_based_on_role(
            user, 
            "New User Created", 
            f"Admin {self.request.user.full_name} has created a new user profile for {user.full_name} ({user.role}).",
            "system",
            sender=self.request.user if self.request.user.is_authenticated else user
        )

    def perform_update(self, serializer):
        instance = self.get_object()
        old_manager = instance.manager
        old_active = instance.is_active
        user = serializer.save()
        
        # 1. Notify on Manager Assignment
        if user.manager and user.manager != old_manager:
            NotificationService.notify_based_on_role(
                user,
                "New Team Member Assigned",
                f"User {user.full_name} has been assigned to your team.",
                "system",
                sender=self.request.user
            )
        
        # 2. Notify on Deactivation
        if old_active and not user.is_active:
            if user.manager:
                NotificationService._send_notifications(
                    self.request.user,
                    [user.manager],
                    "Team Member Deactivated",
                    f"User {user.full_name} has been deactivated by an administrator.",
                    "system"
                )

    def perform_destroy(self, instance):
        manager = instance.manager
        user_name = instance.full_name
        admin_name = self.request.user.full_name
        
        # We notify BEFORE deletion so we still have the object context if needed, 
        # though here we just need the manager reference.
        if manager:
            NotificationService._send_notifications(
                self.request.user,
                [manager],
                "Team Member Removed",
                f"User {user_name} has been permanently removed from the system by Admin {admin_name}.",
                "system"
            )
        instance.delete()

    @action(detail=True, methods=['post'], permission_classes=[IsAdmin])
    def reset_password(self, request, pk=None):
        target_user = self.get_object()
        new_password = request.data.get('new_password')
        if not new_password:
            return Response({'detail': 'new_password required.'}, status=400)
        target_user.set_password(new_password)
        target_user.save()
        AuditService.log(request.user, 'update', f'Password reset for {target_user.email}', request)
        return Response({'detail': 'Password reset successfully.'})


class DepartmentViewSet(viewsets.ModelViewSet):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer

    def get_permissions(self):
        if self.action in ('list', 'retrieve'):
            return [IsAuthenticated()]
        return [IsAdmin()]


# ═══════════════════════════════════════════════════════════════════════════════
# Module 4 & 5: ATTENDANCE ENGINE + WORK SESSIONS
# ═══════════════════════════════════════════════════════════════════════════════

class AttendanceViewSet(viewsets.ReadOnlyModelViewSet):
    """Attendance records — read only. Status is always auto-calculated."""
    serializer_class = AttendanceSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = {
        'date': ['exact', 'gte', 'lte'],
        'status': ['exact', 'in'],
        'is_flagged': ['exact'],
    }
    ordering_fields = ['date', 'created_at']
    ordering = ['-date']

    def get_permissions(self):
        return [IsAuthenticated()]

    def get_queryset(self):
        user = self.request.user
        from django.db.models import Prefetch
        
        # Prefetch with explicit ordering for in-memory serializer speed
        work_sessions_prefetch = Prefetch(
            'work_sessions', 
            queryset=WorkSession.objects.order_by('-start_time').prefetch_related(
                Prefetch('break_sessions', queryset=BreakSession.objects.order_by('-start_time')),
                Prefetch('idle_logs', queryset=IdleLog.objects.order_by('-start_time'))
            )
        )

        qs = Attendance.objects.select_related('user', 'reviewed_by').prefetch_related(work_sessions_prefetch)
        if user.role == 'admin':
            return qs.all()
        if user.role == 'manager':
            subordinate_ids = user.subordinates.values_list('id', flat=True)
            return qs.filter(Q(user=user) | Q(user__in=subordinate_ids))
        return qs.filter(user=user)

    @action(detail=False, methods=['post'])
    def override_status(self, request):
        """Admin override for a user's attendance status."""
        if request.user.role != 'admin':
            return Response({'error': 'Only admins can override attendance.'}, status=403)

        user_id = request.data.get('user_id')
        date_str = request.data.get('date')
        new_status = request.data.get('status')
        remark = request.data.get('remark', 'Status updated by admin')

        if not all([user_id, date_str, new_status]):
            return Response({'error': 'user_id, date, and status are required'}, status=400)

        from .models import User
        try:
            employee = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=404)

        attendance, created = Attendance.objects.get_or_create(
            user=employee,
            date=date_str,
            defaults={
                'status': new_status,
                'manager_remark': remark,
                'is_flagged': True
            }
        )

        if not created:
            attendance.status = new_status
            attendance.manager_remark = remark
            attendance.is_flagged = True
            attendance.save(update_fields=['status', 'manager_remark', 'is_flagged', 'updated_at'])

        # Notify Employee & Manager
        from .services import NotificationService
        NotificationService.notify_attendance_override(request.user, employee, date_str, new_status)

        return Response({
            'status': 'success',
            'message': f'Attendance for {employee.full_name} on {date_str} updated to {new_status}.',
            'attendance_id': attendance.id
        })

    @action(detail=False, methods=['get'])
    def today(self, request):
        """Get current user's today attendance."""
        today = date.today()
        attendance = Attendance.objects.filter(user=request.user, date=today).first()
        if not attendance:
            return Response({'detail': 'No attendance record for today.'}, status=404)
        return Response(AttendanceSerializer(attendance).data)

    @action(detail=False, methods=['get'])
    def todays_absences(self, request):
        """Get today's absences and leaves for employees and managers."""
        today = date.today()
        
        # 1. Get all active employees and managers
        users = User.objects.filter(is_active=True, role__in=['employee', 'manager'])
        
        # 2. Get today's actual attendance records
        attendances = Attendance.objects.filter(
            user__in=users,
            date=today
        ).select_related('user')
        
        att_map = {a.user_id: a for a in attendances}
        
        # 3. Get any valid leave requests spanning today (Approved or Pending)
        leaves = LeaveRequest.objects.filter(
            employee__in=users, 
            from_date__lte=today, 
            to_date__gte=today
        ).exclude(status__in=[LeaveRequest.STATUS_REJECTED, LeaveRequest.STATUS_CANCELLED])
        
        # Map leaves: Prioritize Approved over Pending
        leave_map = {}
        for l in leaves:
            if l.employee_id not in leave_map or l.status == LeaveRequest.STATUS_APPROVED:
                leave_map[l.employee_id] = l
        
        result = []
        for u in users:
            att = att_map.get(u.id)
            leave = leave_map.get(u.id)
            
            # Case 1: Active/Approved/Pending Leave Record
            if leave:
                # If it's approved, the status is 'on_leave'
                # If it's pending, we still want to show it in the leaves table as a justification for absence
                status = 'on_leave' if leave.status == LeaveRequest.STATUS_APPROVED else 'pending_leave'
                result.append({
                    'id': f"leave_{u.id}_{today}",
                    'user_name': u.full_name,
                    'user_email': u.email,
                    'user_role': u.role,
                    'status': status,
                    'leave_type': leave.get_leave_type_display(),
                    'reason': leave.reason or f"Request for {leave.get_leave_type_display()}"
                })
            elif att:
                # Case 2: Attendance record exists
                if att.status == Attendance.STATUS_ON_LEAVE:
                    result.append({
                        'id': str(att.id),
                        'user_name': u.full_name,
                        'user_email': u.email,
                        'user_role': u.role,
                        'status': att.status,
                        'leave_type': 'Leave',
                        'reason': att.manager_remark or 'Approved Leave'
                    })
                elif att.status == Attendance.STATUS_ABSENT:
                    # Explicit unannounced absence
                    result.append({
                        'id': str(att.id),
                        'user_name': u.full_name,
                        'user_email': u.email,
                        'user_role': u.role,
                        'status': att.status,
                        'leave_type': 'None',
                        'reason': 'Unannounced Absence'
                    })
                else:
                    # They are working or present, usually don't show in "Leaves" table 
                    # but if they are here, we might skip or show as 'present'
                    pass 
            else:
                # Case 3: No record at all (Implicit Absent)
                result.append({
                    'id': f"dummy_{u.id}_{today}",
                    'user_name': u.full_name,
                    'user_email': u.email,
                    'user_role': u.role,
                    'status': 'absent',
                    'leave_type': 'None',
                    'reason': 'No Login / Not Checked In'
                })
                
        return Response(result)

    @action(detail=True, methods=['post'], permission_classes=[IsManager])
    def review(self, request, pk=None):
        """Manager reviews and adds remarks to flagged attendance."""
        attendance = self.get_object()
        serializer = AttendanceReviewSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        action = serializer.validated_data.get('action')
        remark = serializer.validated_data.get('manager_remark')

        if action == 'approve':
            attendance.status = Attendance.STATUS_HALF_DAY
            attendance.is_flagged = False
            if not remark:
                remark = "Approved as half-day by manager."
        elif action == 'reject':
            attendance.status = Attendance.STATUS_ABSENT
            attendance.is_flagged = False
            if not remark:
                remark = "Rejected by manager (Absence confirmed)."

        if remark:
            attendance.manager_remark = remark
            
        if 'is_flagged' in serializer.validated_data:
            attendance.is_flagged = serializer.validated_data['is_flagged']

        attendance.reviewed_by = request.user
        attendance.reviewed_at = timezone.now()
        
        attendance.save(update_fields=['status', 'manager_remark', 'is_flagged', 'reviewed_by', 'reviewed_at', 'updated_at'])

        AuditService.log(
            request.user, 'update',
            f'Attendance reviewed ({action}) for {attendance.user.email} on {attendance.date}',
            request
        )
        return Response(AttendanceSerializer(attendance).data)


class WorkSessionView(APIView):
    """Module 5: Start and Stop work sessions."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        """Get today's active work session."""
        today = date.today()
        try:
            attendance = Attendance.objects.filter(user=request.user, date=today).first()
            if not attendance:
                return Response({'active_session': None, 'attendance': None})

            open_session = WorkSession.objects.filter(
                attendance=attendance, end_time__isnull=True
            ).first()
            session_data = WorkSessionSerializer(open_session).data if open_session else None
            attendance_data = AttendanceSerializer(attendance).data
            return Response({'active_session': session_data, 'attendance': attendance_data})
        except Exception as e:
            import logging
            logging.getLogger(__name__).exception("Unexpected error during session fetch")
            return Response({'error': f'System Error: {str(e)}'}, status=500)

    def post(self, request):
        """Start work session."""
        action = request.data.get('action')

        if action == 'start':
            try:
                session, created = WorkSessionService.start_session(request.user, request)
                return Response({
                    'status': 'started' if created else 'already_running',
                    'session': WorkSessionSerializer(session).data,
                })
            except ValueError as e:
                return Response({'error': str(e)}, status=400)
            except Exception as e:
                import logging
                logging.getLogger(__name__).exception('Critical error in start_session')
                return Response({'error': 'Unexpected System Error. Please contact support.'}, status=500)
        elif action == 'stop':
            try:
                session, stopped = WorkSessionService.stop_session(request.user)
                if not stopped:
                    return Response({'detail': 'No active session to stop.'}, status=400)
                today = date.today()
                attendance = Attendance.objects.get(user=request.user, date=today)
                return Response({
                    'status': 'stopped',
                    'session': WorkSessionSerializer(session).data,
                    'attendance': AttendanceSerializer(attendance).data,
                })
            except Exception as e:
                import logging
                logging.getLogger(__name__).exception('Critical error in stop_session')
                return Response({'error': f'Unexpected System Error: {str(e)}'}, status=500)
        else:
            return Response({'detail': 'action must be "start" or "stop".'}, status=400)


# ═══════════════════════════════════════════════════════════════════════════════
# Module 6: BREAK MANAGEMENT
# ═══════════════════════════════════════════════════════════════════════════════

class BreakSessionView(APIView):
    """Start/Stop break within an active work session."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        action = request.data.get('action')

        if action == 'start':
            try:
                break_session, created = BreakSessionService.start_break(request.user)
                return Response({
                    'status': 'started' if created else 'already_on_break',
                    'break_session': BreakSessionSerializer(break_session).data,
                })
            except ValueError as e:
                return Response({'error': str(e)}, status=400)
            except Exception as e:
                return Response({'error': 'System error starting break'}, status=500)
        elif action == 'stop':
            try:
                break_session, stopped = BreakSessionService.stop_break(request.user)
                if not stopped:
                    return Response({'error': 'No active break to stop.'}, status=400)
                return Response({
                    'status': 'stopped',
                    'break_session': BreakSessionSerializer(break_session).data,
                })
            except ValueError as e:
                return Response({'error': str(e)}, status=400)
            except Exception as e:
                return Response({'error': 'System error stopping break'}, status=500)
        else:
            return Response({'error': 'action must be "start" or "stop".'}, status=400)


# ═══════════════════════════════════════════════════════════════════════════════
# Module 7 & 8: IDLE DETECTION + REAL-TIME STATUS
# ═══════════════════════════════════════════════════════════════════════════════

class IdleView(APIView):
    """Frontend fires this when 15 mins of inactivity detected (or when resumed)."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        action_type = request.data.get('action')

        if action_type == 'start':
            try:
                start_time = request.data.get('start_time')
                reason = request.data.get('reason')
                idle_log, created = IdleService.start_idle(request.user, start_time, reason)
                if idle_log is None:
                    return Response({'error': 'No active work session found.'}, status=400)
                return Response({
                    'status': 'idle_started' if created else 'already_idle',
                    'idle_log': IdleLogSerializer(idle_log).data,
                })
            except ValueError as e:
                return Response({'error': str(e)}, status=400)
            except Exception as e:
                return Response({'error': 'System error starting idle log'}, status=500)
        elif action_type == 'stop':
            try:
                idle_log, stopped = IdleService.stop_idle(request.user)
                if not stopped:
                    return Response({'error': 'No active idle period found.'}, status=400)
                return Response({
                    'status': 'resumed',
                    'idle_log': IdleLogSerializer(idle_log).data,
                })
            except ValueError as e:
                return Response({'error': str(e)}, status=400)
            except Exception as e:
                return Response({'error': 'System error stopping idle log'}, status=500)
        else:
            return Response({'error': 'action must be "start" or "stop".'}, status=400)


class SyncSessionView(APIView):
    """
    Heartbeat sync from frontend.
    Updates the current attendance record with the frontend's latest counters.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        user = request.user
        try:
            attendance, _ = AttendanceService.get_or_create_today(user)
            
            # Recalculate based on current logs (which include open sessions now)
            AttendanceService.recalculate_status(attendance)
            
            return Response({
                'status': 'synced',
                'work_seconds': attendance.total_work_seconds,
                'idle_seconds': attendance.total_idle_seconds,
                'live_status': StatusService.get_user_status(user)['status']
            })
        except Exception as e:
            return Response({'error': str(e)}, status=500)




class RealTimeStatusView(APIView):
    """Module 8: Real-time status indicator for dashboard."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        target_id = request.query_params.get('user_id')
        if target_id and request.user.role in ('admin', 'manager'):
            try:
                target_user = User.objects.get(id=target_id)
            except User.DoesNotExist:
                return Response({'detail': 'User not found.'}, status=404)
        else:
            target_user = request.user

        try:
            result = StatusService.get_user_status(target_user)

            # Serialize non-model fields
            attendance_data = AttendanceSerializer(result['attendance']).data if result.get('attendance') else None
            session_data = WorkSessionSerializer(result['session']).data if result.get('session') else None

            return Response({
                'status': result['status'],
                'user_id': str(target_user.id),
                'user_name': target_user.full_name,
                'attendance': attendance_data,
                'session': session_data,
            })
        except Exception as e:
            import logging
            logging.getLogger(__name__).exception(f"Real-time status fetch failed for user {target_user.id}")
            return Response({'error': f"Status Fetch Error: {str(e)}"}, status=500)


class TeamStatusView(APIView):
    """Get real-time status for all team members (manager/admin)."""
    permission_classes = [IsManager]

    def get(self, request):
        user = request.user
        if user.role == 'admin':
            # Admins see Managers and Employees
            team = User.objects.filter(
                is_active=True, 
                role__in=['manager', 'employee']
            ).select_related('department', 'manager')
        elif user.role == 'manager':
            # Managers see their subordinates PLUS themselves (to see their own status on the same page)
            subordinate_ids = list(user.subordinates.values_list('id', flat=True))
            subordinate_ids.append(user.id)
            team = User.objects.filter(
                id__in=subordinate_ids,
                is_active=True
            ).select_related('department', 'manager')
        else:
            team = User.objects.none()

        result = []
        for member in team:
            # Optionally skip self if requested, but generally useful to see self too
            status_data = StatusService.get_user_status(member)
            result.append({
                'user_id': str(member.id),
                'user_name': member.full_name,
                'email': member.email,
                'role': member.role,
                'department': member.department.name if member.department else None,
                'status': status_data['status'],
            })
        return Response(result)


class TeamTimesheetView(APIView):
    """
    Module 8 Extension: Manager-scoped team timesheet.
    Returns today's attendance records for all employees.
    If no attendance exists for an active employee, a dummy 'absent' record is generated.
    """
    permission_classes = [IsManager]

    def get(self, request):
        user = request.user
        
        # Managers should only see their subordinates. Admins see all employees.
        if user.role == 'manager':
            subordinate_ids = list(user.subordinates.values_list('id', flat=True))
            # Include the manager themselves in their team timesheet
            subordinate_ids.append(user.id)
            team = User.objects.filter(id__in=subordinate_ids, is_active=True).select_related('department')
        else:
            # Admins see all staff (Managers and Employees)
            team = User.objects.filter(is_active=True, role__in=['manager', 'employee']).select_related('department')

        # 2. Get today's actual attendance records for these employees
        today = date.today()
        # Optional: frontend could pass ?date=YYYY-MM-DD
        date_str = request.query_params.get('date')
        if date_str:
            try:
                from datetime import datetime
                today = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                pass

        from django.db.models import Prefetch
        
        # Order-aware prefetching for speed
        work_sessions_prefetch = Prefetch(
            'work_sessions', 
            queryset=WorkSession.objects.order_by('-start_time').prefetch_related(
                Prefetch('break_sessions', queryset=BreakSession.objects.order_by('-start_time')),
                Prefetch('idle_logs', queryset=IdleLog.objects.order_by('-start_time'))
            )
        )

        attendances = Attendance.objects.filter(
            user__in=team,
            date=today
        ).select_related('user', 'reviewed_by').prefetch_related(work_sessions_prefetch)
        
        # Map user ID to their attendance record
        att_map = {a.user_id: a for a in attendances}

        enriched = []
        for member in team:
            member_att = att_map.get(member.id)
            
            if member_att:
                # Use actual attendance data (Serializer handles last_logout and hours)
                record = AttendanceSerializer(member_att).data
                record['user_name'] = member.full_name
                record['user_email'] = member.email
                if member.department:
                    record['department_name'] = member.department.name
            else:
                # Generate a dummy 'absent' record
                from .models import AttendancePolicy
                policy = AttendancePolicy.objects.filter(is_active=True, department=member.department).first()
                if not policy:
                    policy = AttendancePolicy.objects.filter(is_active=True, department__isnull=True).first()
                
                target_seconds = float(policy.min_working_hours * 3600) if policy else 28800 # 8h fallback

                record = {
                    'id': f"dummy_{member.id}_{today}",
                    'user': member.id,
                    'user_name': member.full_name,
                    'user_email': member.email,
                    'user_role': member.role,
                    'department_name': member.department.name if member.department else None,
                    'date': str(today),
                    'status': 'absent',
                    'work_hours': '0.00',
                    'total_work_seconds': 0,
                    'total_break_seconds': 0,
                    'total_idle_seconds': 0,
                    'missing_seconds': target_seconds,
                    'is_flagged': False,
                    'flag_reason': '',
                    'manager_remark': '',
                    'reviewed_by': None,
                    'reviewed_at': None,
                    'first_login': None,
                    'last_logout': None
                }

            # Annotate with live status (only makes sense for today, but we fetch it regardless)
            if today == date.today():
                status_data = StatusService.get_user_status(member)
                record['live_status'] = status_data['status']
            else:
                record['live_status'] = 'offline'

            enriched.append(record)

        return Response(enriched)


# ═══════════════════════════════════════════════════════════════════════════════
# Module 10: LEAVE MANAGEMENT
# ═══════════════════════════════════════════════════════════════════════════════

class LeaveRequestViewSet(viewsets.ModelViewSet):
    serializer_class = LeaveRequestSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['status', 'leave_type']
    search_fields = ['employee__email', 'employee__first_name', 'employee__last_name']
    ordering = ['-created_at']

    def get_permissions(self):
        return [IsAuthenticated()]

    def get_queryset(self):
        user = self.request.user
        if user.role == 'admin':
            return LeaveRequest.objects.select_related('employee', 'reviewed_by').all()
        if user.role == 'manager':
            subordinate_ids = user.subordinates.values_list('id', flat=True)
            return LeaveRequest.objects.filter(
                Q(employee=user) | Q(employee__in=subordinate_ids)
            ).select_related('employee', 'reviewed_by')
        return LeaveRequest.objects.filter(employee=user).select_related('employee', 'reviewed_by')

    @action(detail=False, methods=['get'])
    def today(self, request):
        today = date.today()
        # Fetch leaves overlapping with today
        qs = self.get_queryset().filter(from_date__lte=today, to_date__gte=today)
        # Requirement: "employee and manager only leaves" 
        qs = qs.filter(employee__role__in=['employee', 'manager'])
        
        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data)

    def create(self, request, *args, **kwargs):
        serializer = LeaveApplySerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        leave = LeaveService.apply_leave(request.user, serializer.validated_data, request)
        NotificationService.notify_managers_and_admins(
            request.user, "Leave Request", f"{request.user.full_name} applied for leave.", "leave"
        )
        return Response(LeaveRequestSerializer(leave).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'], permission_classes=[IsManager])
    def review(self, request, pk=None):
        leave_request = self.get_object()
        if leave_request.status != LeaveRequest.STATUS_PENDING:
            return Response({'detail': 'Leave is already reviewed.'}, status=400)

        serializer = LeaveReviewSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        updated = LeaveService.review_leave(
            leave_request,
            request.user,
            serializer.validated_data['action'],
            serializer.validated_data.get('comment', ''),
            request,
        )
        return Response(LeaveRequestSerializer(updated).data)

    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        leave_request = self.get_object()
        if leave_request.employee != request.user:
            return Response({'detail': 'Forbidden.'}, status=403)
        if leave_request.status != LeaveRequest.STATUS_PENDING:
            return Response({'detail': 'Cannot cancel a reviewed leave.'}, status=400)
        leave_request.status = LeaveRequest.STATUS_CANCELLED
        leave_request.save()
        return Response(LeaveRequestSerializer(leave_request).data)


# ═══════════════════════════════════════════════════════════════════════════════
# Module 11: HOLIDAY MANAGEMENT
# ═══════════════════════════════════════════════════════════════════════════════

class HolidayViewSet(viewsets.ModelViewSet):
    queryset = Holiday.objects.all().order_by('date')
    serializer_class = HolidaySerializer

    def get_permissions(self):
        if self.action in ('list', 'retrieve'):
            return [IsAuthenticated()]
        return [IsAdmin()]


# ═══════════════════════════════════════════════════════════════════════════════
# Module 12: MANAGER REVIEW
# (Handled via AttendanceViewSet.review action above)
# ═══════════════════════════════════════════════════════════════════════════════

class TeamListView(APIView):
    """
    Simple view to return a list of subordinates for a manager,
    or all employees for an admin. Used for filtering dropdowns.
    """
    permission_classes = [IsManager]

    def get(self, request):
        user = request.user
        if user.role == 'manager':
            team = user.subordinates.filter(is_active=True)
        else:
            # Admins see all productive staff (Managers and Employees)
            team = User.objects.filter(role__in=['manager', 'employee'], is_active=True)
        data = [{'id': m.id, 'full_name': m.full_name, 'email': m.email} for m in team]
        return Response(data)

class FlaggedAttendanceView(generics.ListAPIView):
    """List all flagged attendance records for manager review."""
    serializer_class = AttendanceSerializer
    permission_classes = [IsManager]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ['date', 'status']
    ordering = ['-date']

    def get_queryset(self):
        user = self.request.user
        review_status = self.request.query_params.get('review_status', 'pending')
        target_user_id = self.request.query_params.get('user_id')
        from_date = self.request.query_params.get('from_date')
        to_date = self.request.query_params.get('to_date')

        if review_status == 'reviewed':
            # Records that were flagged but are now reviewed (is_flagged=False, reviewed_by present)
            # OR records that were reviewed and kept as flagged (is_flagged=True, reviewed_by present)
            qs = Attendance.objects.filter(reviewed_by__isnull=False).select_related('user', 'reviewed_by')
        else:
            # Default: only pending (is_flagged=True, not necessarily reviewed yet)
            qs = Attendance.objects.filter(is_flagged=True, reviewed_by__isnull=True).select_related('user', 'reviewed_by')

        if user.role == 'manager':
            subordinate_ids = user.subordinates.values_list('id', flat=True)
            qs = qs.filter(user__in=subordinate_ids)
            
        if target_user_id and target_user_id != 'all':
            qs = qs.filter(user_id=target_user_id)
            
        if from_date:
            qs = qs.filter(date__gte=from_date)
        if to_date:
            qs = qs.filter(date__lte=to_date)
            
        return qs.order_by('-date', '-updated_at')


# ═══════════════════════════════════════════════════════════════════════════════
# Module 13: POLICY CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════

class AttendancePolicyViewSet(viewsets.ModelViewSet):
    queryset = AttendancePolicy.objects.all()
    serializer_class = AttendancePolicySerializer

    def get_permissions(self):
        if self.action in ('list', 'retrieve'):
            return [IsAuthenticated()] # Allow all authenticated to see policy, or at least Manager/Admin. We use IsAuthenticated since Employee might need it for info.
        return [IsAdmin()]

    def perform_create(self, serializer):
        instance = serializer.save()
        
        # Notify all users if a new policy is created (and likely set to active)
        NotificationService.notify_all_active_users(
            self.request.user,
            "New Attendance Policy",
            f"Admin {self.request.user.full_name} has created a new global policy: {instance.name}",
            "system"
        )
        StatusService.broadcast_policy_update()

    def perform_update(self, serializer):
        # Capture old values for comparison
        old_instance = self.get_object()
        
        # Track fields and their human-readable labels
        fields_to_track = {
            'name': 'Policy Name',
            'idle_threshold_minutes': 'Idle Threshold (min)',
            'min_working_hours': 'Full-Day Hours',
            'present_hours': 'Present Hours',
            'half_day_hours': 'Half-Day Hours',
            'shift_start_time': 'Shift Start',
            'shift_end_time': 'Shift End',
            'session_timeout_hours': 'Session Timeout (h)'
        }

        old_values = {field: getattr(old_instance, field) for field in fields_to_track}
        
        # Save the new state
        instance = serializer.save()

        # Build list of specific changes with Old -> New comparison
        changes = []
        changes_dict = {} # For Audit Log extra_data
        
        for field, label in fields_to_track.items():
            old_val = old_values[field]
            new_val = getattr(instance, field)
            if old_val != new_val:
                # Format time objects for readability
                fmt_old = old_val.strftime("%I:%M %p") if hasattr(old_val, 'strftime') else old_val
                fmt_new = new_val.strftime("%I:%M %p") if hasattr(new_val, 'strftime') else new_val
                
                changes.append(f"{label}: {fmt_old} → {fmt_new}")
                changes_dict[field] = {'old': str(old_val), 'new': str(new_val)}

        if not changes:
            return

        # Detailed message for notifications and logs
        msg_header = f"Global Attendance Policy '{instance.name}' Updated:"
        msg_body = "\n".join([f"• {c}" for c in changes])
        full_msg = f"{msg_header}\n{msg_body}"

        AuditService.log(
            self.request.user, 'policy_change',
            f'Policy "{instance.name}" updated by {self.request.user.full_name}',
            self.request,
            extra_data={'changes': changes_dict, 'policy_name': instance.name}
        )

        NotificationService.notify_all_active_users(
            self.request.user,
            "Policy Rule Updated",
            full_msg,
            "system"
        )

        # Trigger background recalculation for recent attendance records to reflect new policy
        try:
            from datetime import date, timedelta
            from .models import Attendance
            from .services import AttendanceService
            
            # Recalculate for the last 3 days to ensure consistency
            recent_dates = [date.today() - timedelta(days=i) for i in range(3)]
            target_attendances = Attendance.objects.filter(date__in=recent_dates)
            for att in target_attendances:
                AttendanceService.recalculate_status(att)
        except Exception as e:
            # Don't fail the update if recalculation fails
            pass


    @action(detail=False, methods=['post'], permission_classes=[IsAdmin])
    def reset_day_sessions(self, request):
        """Administrative action to clear all work sessions and reset attendance for a specific date."""
        date_str = request.data.get('date')
        if not date_str:
            return Response({'error': 'Date is required.'}, status=400)
        
        try:
            from datetime import datetime
            target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=400)

        # 1. Find all attendance for this date
        attendances = Attendance.objects.filter(date=target_date)
        count = attendances.count()

        if count == 0:
            return Response({'message': f'No attendance records found for {date_str}.'})

        # 2. Delete all work sessions associated with these attendance records
        # This will also delete BreakSession and IdleLog via CASCADE
        WorkSession.objects.filter(attendance__in=attendances).delete()

        # 3. Reset attendance status
        holiday = Holiday.objects.filter(date=target_date).first()
        new_status = Attendance.STATUS_HOLIDAY if holiday else Attendance.STATUS_ABSENT
        remark = f"Holiday: {holiday.name}" if holiday else "Attendance reset by administrator."

        attendances.update(
            status=new_status,
            total_work_seconds=0,
            total_break_seconds=0,
            total_idle_seconds=0,
            manager_remark=remark,
            is_flagged=False,
            flag_reason="",
            updated_at=timezone.now()
        )

        # 4. Log action
        AuditService.log(
            request.user, 'update',
            f'Administrative reset of all attendance sessions for {date_str}',
            request,
            extra_data={'date': date_str, 'records_affected': count}
        )

        # 5. Broadcast status change (so users see they are clocked out)
        StatusService.broadcast_policy_update()

        return Response({
            'status': 'success',
            'message': f'Successfully reset {count} attendance sessions for {date_str}.'
        })


# ═══════════════════════════════════════════════════════════════════════════════
# Module 14: REPORTING
# ═══════════════════════════════════════════════════════════════════════════════

class ReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        report_type = request.query_params.get('type', 'summary')
        days = int(request.query_params.get('days', 7))
        user_id = request.query_params.get('user_id')

        # Determine which users to report on for isolation
        user_ids = None  # None means 'all' (Admin only)
        if request.user.role == 'employee':
            user_ids = [request.user.id]
        elif request.user.role == 'manager':
            # Managers can only see their subordinates
            subordinate_ids = list(request.user.subordinates.values_list('id', flat=True))
            if user_id:
                # Check if the requested user_id is actually a subordinate
                # Use string comparison for UUID safety
                subordinate_ids_str = [str(sid) for sid in subordinate_ids]
                if str(user_id) in subordinate_ids_str:
                    user_ids = [user_id]
                else:
                    return Response({'detail': 'Permission denied: User is not your subordinate.'}, status=403)
            else:
                user_ids = subordinate_ids
        elif request.user.role == 'admin':
            if user_id:
                user_ids = [user_id]
            # else user_ids remains None (all)

        # Date range parsing
        from_date = request.query_params.get('from_date')
        to_date = request.query_params.get('to_date')

        if from_date:
            from datetime import datetime
            try:
                from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
            except ValueError:
                from_date = None
        if to_date:
            from datetime import datetime
            try:
                to_date = datetime.strptime(to_date, '%Y-%m-%d').date()
            except ValueError:
                to_date = None

        if report_type == 'summary':
            data = ReportService.get_attendance_summary(user_ids, from_date, to_date)
        elif report_type == 'daily':
            # Update get_daily_data to also support user_ids list
            data = ReportService.get_daily_data(user_ids, days)
        elif report_type == 'team' and request.user.role in ('admin', 'manager'):
            user = request.user
            if user.role == 'manager':
                subordinate_ids = list(user.subordinates.values_list('id', flat=True))
                # Optionally include self in charts? Dashboard usually shows team. We'll stick to subordinates for charts.
            else:
                # Admins see all Managers and Employees in charts, but NOT other admins
                subordinate_ids = list(User.objects.filter(role__in=['manager', 'employee']).values_list('id', flat=True))
            data = []
            for uid in subordinate_ids:
                try:
                    member = User.objects.get(id=uid)
                    summary = ReportService.get_attendance_summary([uid], from_date, to_date)
                    summary['user_id'] = str(uid)
                    summary['user_name'] = member.full_name
                    data.append(summary)
                except User.DoesNotExist:
                    pass
        else:
            return Response({'detail': 'Invalid report type.'}, status=400)

        return Response(data)


# ═══════════════════════════════════════════════════════════════════════════════
# Module 15: AUDIT LOGS
# ═══════════════════════════════════════════════════════════════════════════════

class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = AuditLogSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = {
        'timestamp': ['exact', 'date', 'gte', 'lte'],
        'action_type': ['exact'],
        'user': ['exact']
    }
    search_fields = ['description', 'user__email']
    ordering = ['-timestamp']

    def get_queryset(self):
        user = self.request.user
        qs = AuditLog.objects.select_related('user').all()
        if user.role == 'admin':
            return qs
        if user.role == 'manager':
            subordinate_ids = user.subordinates.values_list('id', flat=True)
            return qs.filter(Q(user=user) | Q(user_id__in=subordinate_ids))
        return qs.filter(user=user)


# ═══════════════════════════════════════════════════════════════════════════════
# Module 16: EXPORT
# ═══════════════════════════════════════════════════════════════════════════════

class ExportView(APIView):
    permission_classes = [IsManager] # Admin/Manager can export

    def get(self, request):
        from datetime import date, datetime, time
        export_type = request.query_params.get('type', 'attendance')
        file_format = request.query_params.get('export_format', 'csv')
        
        # Support both 'from_date' and 'date__gte' for frontend consistency
        raw_from = request.query_params.get('from_date') or request.query_params.get('date__gte')
        raw_to = request.query_params.get('to_date') or request.query_params.get('date__lte')
        
        # Robust Datetime Range Parsing: 
        # Create full datetime objects for the start and end of the specified range.
        # This is more reliable than __date filtering on some database engines.
        start_dt = None
        end_dt = None
        try:
            if raw_from: 
                start_dt = datetime.combine(date.fromisoformat(raw_from), time.min)
            if raw_to: 
                end_dt = datetime.combine(date.fromisoformat(raw_to), time.max)
        except (ValueError, TypeError):
            pass

        AuditService.log(request.user, 'export', f'Exported {export_type} as {file_format} (Range: {raw_from} to {raw_to})', request)

        if export_type == 'attendance':
            return self._export_attendance(request, start_dt, end_dt, file_format)
        elif export_type == 'leave':
            # Note: _export_leave handles its own parsing internally for gap detection,
            # but we pass raw strings or objects if needed.
            return self._export_leave(request, raw_from, raw_to, file_format)
        elif export_type == 'audit':
            return self._export_audit(request, start_dt, end_dt, file_format)
        elif export_type == 'payroll':
            return self._export_payroll(request, file_format)
        elif export_type == 'holiday':
            return self._export_holiday(request, file_format)
        else:
            return Response({'detail': 'Invalid export type.'}, status=400)

    def _export_attendance(self, request, start_dt, end_dt, file_format):
        from django.http import HttpResponse
        qs = Attendance.objects.select_related('user', 'user__department').all()
        if request.user.role == 'manager':
            subordinate_ids = list(request.user.subordinates.values_list('id', flat=True))
            # Managers should see their own data + subordinates in exports
            qs = qs.filter(Q(user=request.user) | Q(user_id__in=subordinate_ids))

        category = request.query_params.get('category')
        user_id = request.query_params.get('user_id')

        if category == 'employees':
            qs = qs.filter(user__role='employee')
        elif category == 'managers':
            qs = qs.filter(user__role='manager')
        elif category == 'particular_employee' and user_id:
            qs = qs.filter(user_id=user_id)

        # Filters using date or full datetime for precision
        if start_dt:
            qs = qs.filter(date__gte=start_dt.date())
        if end_dt:
            qs = qs.filter(date__lte=end_dt.date())

        if file_format == 'xlsx':
            return self._to_xlsx(
                qs,
                ['Employee', 'Email', 'Department', 'Date', 'Status',
                 'Work Hours', 'Break Hours', 'Idle Hours', 'Productive Hours', 'Flagged'],
                lambda a: [
                    a.user.full_name, a.user.email,
                    a.user.department.name if a.user.department else '',
                    str(a.date), a.status,
                    round(a.total_work_seconds / 3600, 2),
                    round(a.total_break_seconds / 3600, 2),
                    round(a.total_idle_seconds / 3600, 2),
                    round(a.effective_work_seconds / 3600, 2),
                    'Yes' if a.is_flagged else 'No',
                ],
                'attendance_export.xlsx'
            )

        # CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="attendance_export.csv"'
        writer = csv.writer(response)
        writer.writerow(['Employee', 'Email', 'Department', 'Date', 'Status',
                         'Work Hours', 'Break Hours', 'Idle Hours', 'Productive Hours', 'Flagged'])
        for a in qs:
            writer.writerow([
                a.user.full_name, a.user.email,
                a.user.department.name if a.user.department else '',
                str(a.date), a.status,
                round(a.total_work_seconds / 3600, 2),
                round(a.total_break_seconds / 3600, 2),
                round(a.total_idle_seconds / 3600, 2),
                round(a.effective_work_seconds / 3600, 2),
                'Yes' if a.is_flagged else 'No',
            ])
        return response

    def _export_leave(self, request, from_date, to_date, file_format):
        """
        Export leave data including formal requests AND attendance-based absences.
        This ensures the export matches what admins see in the 'Today's Leaves' table.
        """
        from django.http import HttpResponse
        from .models import User, LeaveRequest, Attendance, Holiday
        from datetime import timedelta, date, datetime

        # 1. Standardize date range
        try:
            start_date = date.fromisoformat(from_date) if from_date else date.today() - timedelta(days=7)
            end_date = date.fromisoformat(to_date) if to_date else date.today()
        except ValueError:
            start_date = date.today() - timedelta(days=7)
            end_date = date.today()

        # 2. Filter users based on category
        category = request.query_params.get('category')
        user_id = request.query_params.get('user_id')
        
        users = User.objects.filter(is_active=True).exclude(role='admin')
        if request.user.role == 'manager':
            subordinate_ids = list(request.user.subordinates.values_list('id', flat=True))
            users = users.filter(Q(id=request.user.id) | Q(id__in=subordinate_ids))
            
        if category == 'employees':
            users = users.filter(role='employee')
        elif category == 'managers':
            users = users.filter(role='manager')
        elif category == 'particular_employee' and user_id:
            users = users.filter(id=user_id)
            
        users = users.distinct()

        # 3. Pre-fetch all necessary data to avoid N+1 queries during loop
        leaves = LeaveRequest.objects.filter(
            employee__in=users,
            from_date__lte=end_date,
            to_date__gte=start_date
        ).select_related('employee', 'reviewed_by')
        
        attendances = Attendance.objects.filter(
            user__in=users,
            date__gte=start_date,
            date__lte=end_date
        ).select_related('user')
        
        holidays = Holiday.objects.filter(date__gte=start_date, date__lte=end_date)
        holiday_dates = {h.date: h.name for h in holidays}
        
        # Index data for fast O(1) lookups
        att_map = {(a.user_id, a.date): a for a in attendances}
        leaf_map = {}
        for l in leaves:
            if l.employee_id not in leaf_map: leaf_map[l.employee_id] = []
            leaf_map[l.employee_id].append(l)

        report_rows = []
        today = date.today()

        # 4. Generate rows date-by-date (Gap Detection)
        curr = start_date
        while curr <= end_date:
            is_holiday = curr in holiday_dates
            for u in users:
                # Priority 1: Formal Leave Request
                user_leaves = leaf_map.get(u.id, [])
                matching_leaf = next((l for l in user_leaves if l.from_date <= curr <= l.to_date), None)
                
                if matching_leaf:
                    report_rows.append({
                        'date': curr,
                        'name': u.full_name,
                        'email': u.email,
                        'type': matching_leaf.leave_type,
                        'from': matching_leaf.from_date,
                        'to': matching_leaf.to_date,
                        'days': matching_leaf.duration_days,
                        'status': matching_leaf.status.upper(),
                        'reason': matching_leaf.reason,
                        'reviewed_by': matching_leaf.reviewed_by.full_name if matching_leaf.reviewed_by else '',
                    })
                    continue
                
                # Priority 2: System-detected Absence (Not present and not on holiday)
                if not is_holiday:
                    att = att_map.get((u.id, curr))
                    # NO record usually means absent. (Unless we want to exclude future dates)
                    # We only report absences for today or past days.
                    if curr <= today:
                        is_absent = False
                        reason = "No attendance record found"
                        
                        if not att:
                            is_absent = True
                        elif att.status == 'absent':
                            is_absent = True
                            reason = att.manager_remark or "Marked absent by system"

                        if is_absent:
                            report_rows.append({
                                'date': curr,
                                'name': u.full_name,
                                'email': u.email,
                                'type': 'Unplanned Absence',
                                'from': curr,
                                'to': curr,
                                'days': 1,
                                'status': 'ABSENT',
                                'reason': reason,
                                'reviewed_by': 'System',
                            })
            curr += timedelta(days=1)

        # Sort: Latest date first, then by name
        report_rows.sort(key=lambda x: (x['date'], x['name']), reverse=True)

        if file_format == 'xlsx':
            return self._to_xlsx(
                report_rows,
                ['Date', 'Employee', 'Email', 'Type', 'From', 'To', 'Days', 'Status', 'Reason', 'Reviewed By'],
                lambda row: [
                    str(row['date']), row['name'], row['email'], row['type'],
                    str(row['from']), str(row['to']), row['days'], row['status'],
                    row['reason'], row['reviewed_by']
                ],
                'leave_export.xlsx'
            )

        # CSV Export
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="leave_export.csv"'
        writer = csv.writer(response)
        writer.writerow(['Date', 'Employee', 'Email', 'Type', 'From', 'To', 'Days', 'Status', 'Reason', 'Reviewed By'])
        for row in report_rows:
            writer.writerow([
                str(row['date']), row['name'], row['email'], row['type'],
                str(row['from']), str(row['to']), row['days'], row['status'],
                row['reason'], row['reviewed_by']
            ])
        return response

    def _export_holiday(self, request, file_format):
        from django.http import HttpResponse
        qs = Holiday.objects.all().order_by('date')
        
        if file_format == 'xlsx':
            return self._to_xlsx(
                qs,
                ['Name', 'Date', 'Description'],
                lambda h: [h.name, str(h.date), h.description],
                'holidays_export.xlsx'
            )

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="holidays_export.csv"'
        writer = csv.writer(response)
        writer.writerow(['Name', 'Date', 'Description'])
        for h in qs:
            writer.writerow([h.name, str(h.date), h.description])
        return response

    def _export_audit(self, request, start_dt, end_dt, file_format):
        import csv
        from django.http import HttpResponse
        qs = AuditLog.objects.select_related('user').all().order_by('-timestamp')

        if request.user.role == 'manager':
            subordinate_ids = request.user.subordinates.values_list('id', flat=True)
            qs = qs.filter(Q(user=request.user) | Q(user_id__in=subordinate_ids))

        category = request.query_params.get('category')
        user_id = request.query_params.get('user_id')

        if category == 'employees':
            qs = qs.filter(user__role='employee')
        elif category == 'managers':
            qs = qs.filter(user__role='manager')
        elif category == 'particular_employee' and user_id:
            qs = qs.filter(user_id=user_id)
            
        if start_dt:
            qs = qs.filter(timestamp__gte=start_dt)
        if end_dt:
            qs = qs.filter(timestamp__lte=end_dt)

        if file_format == 'xlsx':
            return self._to_xlsx(
                qs,
                ['User', 'Action', 'Description', 'IP', 'Timestamp'],
                lambda al: [
                    al.user.email if al.user else '',
                    al.action_type, al.description,
                    al.ip_address or '', str(al.timestamp),
                ],
                'audit_export.xlsx'
            )

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="audit_export.csv"'
        writer = csv.writer(response)
        writer.writerow(['User', 'Action', 'Description', 'IP', 'Timestamp'])
        for al in qs:
            writer.writerow([
                al.user.email if al.user else '',
                al.action_type, al.description,
                al.ip_address or '', str(al.timestamp),
            ])
        return response

    def _export_payroll(self, request, file_format):
        from .services import PayrollPrepService
        from django.http import HttpResponse
        
        # Default to current month if not provided
        from datetime import date
        today = date.today()
        month = int(request.query_params.get('month', today.month))
        year = int(request.query_params.get('year', today.year))
        
        payroll_data = PayrollPrepService.prepare_monthly_payroll(month, year)
        
        if file_format == 'xlsx':
            return self._to_xlsx(
                payroll_data,
                ['Employee ID', 'Email', 'Name', 'Department', 'Role', 'Payable Days', 'Present Days', 'Paid Leave Days', 'Reimbursable Expenses'],
                lambda p: [
                    p['employee_id'], p['email'], p['name'], p['department'],
                    p['role'], p['payable_days'], p['present_days'],
                    p['paid_leave_days'], p['reimbursable_expenses']
                ],
                f'payroll_{year}_{month:02d}.xlsx'
            )
            
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="payroll_{year}_{month:02d}.csv"'
        writer = csv.writer(response)
        writer.writerow(['Employee ID', 'Email', 'Name', 'Department', 'Role', 'Payable Days', 'Present Days', 'Paid Leave Days', 'Reimbursable Expenses'])
        for p in payroll_data:
            writer.writerow([
                p['employee_id'], p['email'], p['name'], p['department'],
                p['role'], p['payable_days'], p['present_days'],
                p['paid_leave_days'], p['reimbursable_expenses']
            ])
        return response

    def _to_xlsx(self, queryset, headers, row_fn, filename):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'REMS Export'

        # Header styling
        header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row_num, obj in enumerate(queryset, 2):
            for col, value in enumerate(row_fn(obj), 1):
                ws.cell(row=row_num, column=col, value=value)

        # Auto-width
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

# ═══════════════════════════════════════════════════════════════════════════════
# Phase 10 Expanded Viewsets
# ═══════════════════════════════════════════════════════════════════════════════

from .models import IPWhitelist, Shift, Project, Task, AppUsageLog, Alert, Document, Expense
from .serializers import (
    IPWhitelistSerializer, ShiftSerializer, ProjectSerializer, TaskSerializer,
    AppUsageLogSerializer, AlertSerializer, DocumentSerializer, ExpenseSerializer
)

class IPWhitelistViewSet(viewsets.ModelViewSet):
    queryset = IPWhitelist.objects.all()
    serializer_class = IPWhitelistSerializer
    permission_classes = [IsAuthenticated, IsAdmin]

class ShiftViewSet(viewsets.ModelViewSet):
    queryset = Shift.objects.all()
    serializer_class = ShiftSerializer
    permission_classes = [IsAuthenticated]

class ProjectViewSet(viewsets.ModelViewSet):
    queryset = Project.objects.all()
    serializer_class = ProjectSerializer
    permission_classes = [IsAuthenticated]

class TaskViewSet(viewsets.ModelViewSet):
    serializer_class = TaskSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['status', 'project', 'assigned_to']
    search_fields = ['title', 'description']

    def get_queryset(self):
        user = self.request.user
        if user.role == "admin":
            return Task.objects.all()
        elif user.role == "manager":
            return Task.objects.all()
        return Task.objects.filter(assigned_to=user)

class AppUsageLogViewSet(viewsets.ModelViewSet):
    serializer_class = AppUsageLogSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = {
        'timestamp': ['exact', 'date', 'gte', 'lte']
    }
    ordering = ['-timestamp']

    def get_queryset(self):
        user = self.request.user
        if user.role in ["admin", "manager"]:
            return AppUsageLog.objects.all()
        return AppUsageLog.objects.filter(user=user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

class AlertViewSet(viewsets.ModelViewSet):
    serializer_class = AlertSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Alert.objects.filter(user=self.request.user)

class DocumentViewSet(viewsets.ModelViewSet):
    serializer_class = DocumentSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.role in ["admin", "manager"]:
            return Document.objects.all()
        from django.db.models import Q
        return Document.objects.filter(Q(uploaded_by=user) | Q(is_public=True))

    def perform_create(self, serializer):
        serializer.save(uploaded_by=self.request.user)

class ExpenseViewSet(viewsets.ModelViewSet):
    serializer_class = ExpenseSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.role in ["admin", "manager"]:
            return Expense.objects.all()
        return Expense.objects.filter(user=user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


# ═══════════════════════════════════════════════════════════════════════════════
# Notifications API
# ═══════════════════════════════════════════════════════════════════════════════
from rest_framework.decorators import action

class NotificationViewSet(viewsets.ModelViewSet):
    serializer_class = NotificationSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ['is_read', 'type']
    ordering = ['-created_at']

    def get_queryset(self):
        return Notification.objects.filter(recipient=self.request.user)

    @action(detail=False, methods=['post'])
    def mark_all_read(self, request):
        self.get_queryset().update(is_read=True)
        return Response({'status': 'All notifications marked as read'})

    @action(detail=True, methods=['post'])
    def mark_read(self, request, pk=None):
        notification = self.get_object()
        notification.is_read = True
        notification.save()
        return Response({'status': 'Notification marked as read'})


class ScreenCaptureViewSet(viewsets.ModelViewSet):
    queryset = ScreenCapture.objects.all()
    serializer_class = ScreenCaptureSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        qs = ScreenCapture.objects.select_related('user').all()
        
        if user.role == 'admin':
            return qs
        if user.role == 'manager':
            # Managers can see their own + subordinates
            subordinate_ids = user.subordinates.values_list('id', flat=True)
            return qs.filter(Q(user=user) | Q(user_id__in=subordinate_ids))
        
        # Employees can only see their own (though they usually only upload)
        return qs.filter(user=user)

    def perform_create(self, serializer):
        # Automatically assign the uploading user
        serializer.save(user=self.request.user)

    @action(detail=False, methods=['get'], url_path='latest')
    def latest(self, request):
        """Get the most recent screenshot for each active employee (Admin/Manager only)."""
        if request.user.role not in ['admin', 'manager']:
            return Response({'detail': 'Unauthorized.'}, status=403)

        # Get the latest capture for each user in the visible scope (subordinates/all)
        queryset = self.get_queryset()
        
        # MySQL doesn't support LIMIT in IN subqueries. 
        # We'll use a more compatible approach: 
        # 1. Get the latest timestamp for each user.
        # 2. Fetch the records matching those user/timestamp pairs.
        from django.db.models import Max
        latest_stamps = queryset.values('user').annotate(max_ts=Max('timestamp'))
        
        # Build a filter for the latest records
        if not latest_stamps:
            return Response([])

        # Filter the queryset to only include those exact user/timestamp matches
        # Note: In rare cases of exact same timestamp, this might return multiple, 
        # but for screenshots it's highly unlikely and acceptable.
        from django.db.models import Q
        filter_q = Q()
        for entry in latest_stamps:
            filter_q |= Q(user_id=entry['user'], timestamp=entry['max_ts'])
        
        latest_captures = ScreenCapture.objects.filter(filter_q).select_related('user')

        serializer = self.get_serializer(latest_captures, many=True)
        return Response(serializer.data)

from django.core.management import call_command
import traceback

class DebugDBView(APIView):
    permission_classes = [AllowAny]
    def get(self, request):
        try:
            print("RUNNING ON-DEMAND MIGRATIONS VIA API")
            call_command('migrate', '--noinput')
            return Response({'status': 'Migration successful!'})
        except Exception as e:
            return Response({
                'status': 'Migration failed',
                'error': str(e),
                'traceback': traceback.format_exc()
            }, status=500)

